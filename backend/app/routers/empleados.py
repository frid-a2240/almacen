from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import Empleado, MovimientoResguardo
from app.schemas.empleado import EmpleadoOut, EmpleadoCreate, EmpleadoUpdate
from app.schemas.movimiento_resguardo import MovimientoOut
from app.services.uploads import guardar_archivo
from app.services.resguardo import resguardo_actual_de
from app.deps_auth import usuario_actual

router = APIRouter(prefix="/empleados", tags=["Empleados"], dependencies=[Depends(usuario_actual)])


def _to_out(emp: Empleado) -> EmpleadoOut:
    out = EmpleadoOut.model_validate(emp)
    out.departamento_nombre = emp.departamento_ref.departamento if emp.departamento_ref else None
    return out


@router.get("/", response_model=list[EmpleadoOut])
def listar(db: Session = Depends(get_db)):
    empleados = (
        db.query(Empleado)
        .options(joinedload(Empleado.departamento_ref))
        .order_by(Empleado.nombre_de_empleado)
        .all()
    )
    return [_to_out(e) for e in empleados]


@router.get("/{id_numero_empleado}", response_model=EmpleadoOut)
def obtener(id_numero_empleado: str, db: Session = Depends(get_db)):
    emp = db.get(Empleado, id_numero_empleado)
    if not emp:
        raise HTTPException(404, "Empleado no encontrado")
    return _to_out(emp)


@router.post("/", response_model=EmpleadoOut, status_code=201)
def crear(datos: EmpleadoCreate, db: Session = Depends(get_db)):
    if db.get(Empleado, datos.id_numero_empleado):
        raise HTTPException(409, "Ya existe un empleado con ese número")
    emp = Empleado(**datos.model_dump())
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return _to_out(emp)


@router.put("/{id_numero_empleado}", response_model=EmpleadoOut)
def actualizar(id_numero_empleado: str, datos: EmpleadoUpdate, db: Session = Depends(get_db)):
    emp = db.get(Empleado, id_numero_empleado)
    if not emp:
        raise HTTPException(404, "Empleado no encontrado")
    for campo, valor in datos.model_dump(exclude_unset=True).items():
        setattr(emp, campo, valor)
    db.commit()
    db.refresh(emp)
    return _to_out(emp)


@router.delete("/{id_numero_empleado}", status_code=204)
def eliminar(id_numero_empleado: str, db: Session = Depends(get_db)):
    emp = db.get(Empleado, id_numero_empleado)
    if not emp:
        raise HTTPException(404, "Empleado no encontrado")
    db.delete(emp)
    db.commit()


@router.get("/{id_numero_empleado}/movimientos", response_model=list[MovimientoOut])
def movimientos_de_empleado(id_numero_empleado: str, db: Session = Depends(get_db)):
    return (
        db.query(MovimientoResguardo)
        .options(joinedload(MovimientoResguardo.producto_ref), joinedload(MovimientoResguardo.empleado_ref))
        .filter(MovimientoResguardo.empleado_id == id_numero_empleado)
        .order_by(MovimientoResguardo.fecha_movimiento.desc())
        .all()
    )



# Mismas columnas/orden que la plantilla de AppSheet "IMPRESION INVENTARIO
# APPSHEET MCR.xlsm" (hoja "INVENTARIO PERSONAL"). Los anchos también son los
# de esa plantilla, salvo la columna S (FIRMA...) que ahí quedó con un ancho
# de 255 — claramente un error de arrastre en el original, aquí se deja en
# uno razonable. Las columnas de foto/firma (C, Q, R, S) quedan vacías, tal
# como estaban en la plantilla (angostas/sin usar, no traía fotos incrustadas).
_COLUMNAS_INVENTARIO = [
    "FECHA ADQ.", "# VALE", "FOTO VALE DE SALIDA", "TIPO MOVIMIENTO",
    "ID NUMERO EMPLEADO", "NOMBRE DE EMPLEADO", "PUESTO / POSICION", "DEPARTAMENTO",
    "JEFE INMEDIATO", "STATUS", "CODIGO SAI SKU", "DESCRIPCION", "UDM", "#ECONOM",
    "CLASE / FAMILIA", "QTY", "FOTO PRODUCTO", "FOTO # NUMERO SERIE",
    "FIRMA DE RECIBIDO Y CONFORMIDAD", "OBSERVACIONES", "Costo Unitario", "Costo Total",
]
_ANCHOS_INVENTARIO = [12, 8.9, 0.3, 22, 24.4, 0.1, 20.9, 18.1, 20, 10, 18.4, 49.7, 7.9, 12.6, 17.7, 6.6, 1, 23.7, 40, 18.1, 12.7, 13.7]
_FILA_ENCABEZADO = 7
_FILA_PRIMER_DATO = 8


@router.get("/{id_numero_empleado}/resguardo-excel")
def resguardo_excel(id_numero_empleado: str, db: Session = Depends(get_db)):
    emp = db.get(Empleado, id_numero_empleado)
    if not emp:
        raise HTTPException(404, "Empleado no encontrado")

    filas = resguardo_actual_de(db, id_numero_empleado)
    depto = emp.departamento_ref.departamento if emp.departamento_ref else ""

    wb = Workbook()
    ws = wb.active
    ws.title = "INVENTARIO PERSONAL"

    negrita = Font(bold=True)

    ws["K1"] = "INVENTARIO DE HERRAMIENTA"
    ws["K1"].font = Font(bold=True, size=14)
    ws["M1"] = "FECHA:"
    ws["M1"].font = negrita
    ws["N1"] = date.today()
    ws["N1"].number_format = "dd/mm/yyyy"

    ws["K2"] = "NOMBRE:"
    ws["K2"].font = negrita
    ws["L2"] = emp.nombre_de_empleado
    ws["K3"] = "# EMPLEADO"
    ws["K3"].font = negrita
    ws["L3"] = emp.id_numero_empleado
    ws["U3"] = "Total :"
    ws["U3"].font = negrita
    ws["K4"] = "PUESTO:"
    ws["K4"].font = negrita
    ws["L4"] = emp.puesto_posicion
    ws["M4"] = "___________________________________________"
    ws["K5"] = "ESTATUS:"
    ws["K5"].font = negrita
    ws["L5"] = emp.status_empleado
    ws["M5"] = "FIRMA DE CONFORMIDAD"

    for i, nombre in enumerate(_COLUMNAS_INVENTARIO, start=1):
        celda = ws.cell(row=_FILA_ENCABEZADO, column=i, value=nombre)
        celda.font = negrita

    for offset, f in enumerate(filas):
        r = _FILA_PRIMER_DATO + offset
        ws.cell(row=r, column=1, value=f["fecha"]).number_format = "dd/mm/yyyy"
        ws.cell(row=r, column=2, value=f["numero_de_vale"])
        # C (FOTO VALE DE SALIDA), Q (FOTO PRODUCTO), R (FOTO # NUMERO SERIE),
        # S (FIRMA) quedan vacías — ver nota arriba.
        ws.cell(row=r, column=4, value=f["tipo_movimiento"])
        ws.cell(row=r, column=5, value=emp.id_numero_empleado)
        ws.cell(row=r, column=6, value=emp.nombre_de_empleado)
        ws.cell(row=r, column=7, value=emp.puesto_posicion)
        ws.cell(row=r, column=8, value=depto)
        ws.cell(row=r, column=9, value=emp.jefe_inmediato)
        ws.cell(row=r, column=10, value=emp.status_empleado)
        ws.cell(row=r, column=11, value=f["sku"])
        ws.cell(row=r, column=12, value=f["descripcion"])
        ws.cell(row=r, column=13, value=f["udm"])
        ws.cell(row=r, column=14, value=f["numero_economico"])
        ws.cell(row=r, column=15, value=f["clase_familia"])
        ws.cell(row=r, column=16, value=float(f["cantidad"]))
        ws.cell(row=r, column=20, value=f["observaciones"])
        costo = float(f["costo_unitario"]) if f["costo_unitario"] is not None else None
        ws.cell(row=r, column=21, value=costo)
        if costo is not None:
            ws.cell(row=r, column=22, value=f"=U{r}*P{r}")

    if filas:
        ultima_fila = _FILA_PRIMER_DATO + len(filas) - 1
        ws["U4"] = f"=SUM(V{_FILA_PRIMER_DATO}:V{ultima_fila})"
    else:
        ws["U4"] = 0

    for i, ancho in enumerate(_ANCHOS_INVENTARIO, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"inventario_{id_numero_empleado}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )


@router.post("/{id_numero_empleado}/foto", response_model=EmpleadoOut)
def subir_foto(id_numero_empleado: str, archivo: UploadFile = File(...), db: Session = Depends(get_db)):
    emp = db.get(Empleado, id_numero_empleado)
    if not emp:
        raise HTTPException(404, "Empleado no encontrado")
    emp.foto_empleado = guardar_archivo(archivo, "EMPLEADOS", id_numero_empleado, "FOTO_EMPLEADO")
    db.commit()
    db.refresh(emp)
    return _to_out(emp)
